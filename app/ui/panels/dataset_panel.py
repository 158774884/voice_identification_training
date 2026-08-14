"""
Dataset Manager Panel — import, preview, clean, and split audio datasets.
"""
import os
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QPushButton,
    QTableView, QHeaderView, QTreeView, QFileSystemModel,
    QSplitter, QLabel, QSlider, QSpinBox, QDoubleSpinBox,
    QAbstractItemView, QMessageBox, QFileDialog,
)
from PySide6.QtCore import Qt, QAbstractTableModel, QModelIndex, Signal, Slot
from PySide6.QtGui import QColor

from app.controllers.dataset_controller import DatasetController
from app.models.dataset_model import AudioFileInfo, DatasetModel
from app.ui.widgets.audio_waveform import AudioWaveformWidget
from app.utils.logger import LogManager


class FileTableModel(QAbstractTableModel):
    """Table model for audio file metadata."""

    COLUMNS = ["文件名", "时长(s)", "文本/命令", "说话人", "性别", "年龄", "地区", "状态", "划分"]

    STATUS_COLORS = {
        "valid": QColor("#28a745"),
        "short": QColor("#e6a817"),
        "long": QColor("#e6a817"),
        "silent": QColor("#e6a817"),
        "error": QColor("#dc3545"),
    }

    def __init__(self, parent=None):
        super().__init__(parent)
        self._files: list[AudioFileInfo] = []

    def set_files(self, files: list[AudioFileInfo]):
        self.beginResetModel()
        self._files = files
        self.endResetModel()

    def rowCount(self, parent=QModelIndex()):
        return len(self._files)

    def columnCount(self, parent=QModelIndex()):
        return len(self.COLUMNS)

    def data(self, index: QModelIndex, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        f = self._files[index.row()]

        if role == Qt.DisplayRole:
            col = index.column()
            if col == 0:
                return f.filename
            elif col == 1:
                return f"{f.duration:.1f}"
            elif col == 2:
                return f.text or "-"
            elif col == 3:
                return f.speaker_id or "-"
            elif col == 4:
                return f.gender or "-"
            elif col == 5:
                return f.age or "-"
            elif col == 6:
                return f.region or "-"
            elif col == 7:
                return f.status
            elif col == 8:
                return f.split

        if role == Qt.ForegroundRole and index.column() == 7:
            return self.STATUS_COLORS.get(f.status, QColor("#6c757d"))

        if role == Qt.UserRole:
            return f

        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role == Qt.DisplayRole and orientation == Qt.Horizontal:
            return self.COLUMNS[section]
        return None

    @property
    def files(self):
        return self._files


class DatasetPanel(QWidget):
    """Main dataset management panel."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._controller = DatasetController(self)
        self._log = LogManager()
        self._current_preview_path: str | None = None
        self._setup_ui()
        self._connect_signals()

    def _setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(8, 8, 8, 8)

        # === Top: Import Section ===
        import_group = QGroupBox("数据集导入")
        import_layout = QHBoxLayout(import_group)

        self._import_btn = QPushButton("📂 导入文件夹")
        import_layout.addWidget(self._import_btn)

        self._import_recursive_btn = QPushButton("📂 递归导入 (含子文件夹)")
        import_layout.addWidget(self._import_recursive_btn)

        import_layout.addStretch()

        self._file_count_label = QLabel("已导入: 0 个文件")
        self._file_count_label.setStyleSheet("color: #5f6368; font-weight: bold;")
        import_layout.addWidget(self._file_count_label)

        export_jsonl_btn = QPushButton("导出 JSONL")
        export_jsonl_btn.setObjectName("secondaryBtn")
        import_layout.addWidget(export_jsonl_btn)

        import_xlsx_btn = QPushButton("导入标注(xlsx)")
        import_xlsx_btn.setToolTip("从 xlsx 文件导入文本标注，匹配到已导入的音频文件")
        import_layout.addWidget(import_xlsx_btn)

        prepare_training_btn = QPushButton("准备训练数据")
        prepare_training_btn.setToolTip("将当前数据集导出为训练用 JSONL，并设为训练面板的数据根目录")
        import_layout.addWidget(prepare_training_btn)

        clear_btn = QPushButton("清空")
        clear_btn.setObjectName("dangerBtn")
        import_layout.addWidget(clear_btn)

        main_layout.addWidget(import_group)

        # === Middle: File Table + Waveform ===
        splitter = QSplitter(Qt.Horizontal)

        # --- Left: File Table ---
        self._table = QTableView()
        self._model = FileTableModel(self)
        self._table.setModel(self._model)
        self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._table.setAlternatingRowColors(True)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.setSortingEnabled(True)
        self._table.setColumnWidth(0, 220)  # filename
        self._table.setColumnWidth(1, 60)   # duration
        self._table.setColumnWidth(2, 100)  # text
        self._table.setColumnWidth(7, 60)   # status
        self._table.setColumnWidth(8, 60)   # split
        self._table.verticalHeader().setDefaultSectionSize(24)
        self._table.verticalHeader().setVisible(False)
        splitter.addWidget(self._table)

        # --- Right: Waveform Preview ---
        preview_widget = QWidget()
        preview_layout = QVBoxLayout(preview_widget)

        preview_header = QHBoxLayout()
        preview_label = QLabel("音频预览")
        preview_label.setStyleSheet("font-weight: bold; color: #5f6368;")
        preview_header.addWidget(preview_label)
        preview_header.addStretch()

        self._preview_file_label = QLabel("选择文件以预览")
        self._preview_file_label.setStyleSheet("color: #9aa0a6;")
        preview_header.addWidget(self._preview_file_label)
        preview_layout.addLayout(preview_header)

        self._waveform = AudioWaveformWidget(self)
        preview_layout.addWidget(self._waveform)

        # Playback controls
        playback_layout = QHBoxLayout()
        self._play_btn = QPushButton("▶ 播放")
        self._play_btn.setMaximumWidth(100)
        self._play_btn.setEnabled(False)
        playback_layout.addWidget(self._play_btn)

        self._stop_btn = QPushButton("⏹ 停止")
        self._stop_btn.setMaximumWidth(100)
        self._stop_btn.setEnabled(False)
        self._stop_btn.setObjectName("dangerBtn")
        playback_layout.addWidget(self._stop_btn)

        playback_layout.addStretch()

        preview_layout.addLayout(playback_layout)
        preview_layout.addStretch()

        splitter.addWidget(preview_widget)
        splitter.setSizes([700, 400])
        main_layout.addWidget(splitter)

        # === Bottom: Clean & Split ===
        bottom_group = QGroupBox("数据清洗与划分")
        bottom_layout = QHBoxLayout(bottom_group)

        # Duration filter
        bottom_layout.addWidget(QLabel("最小时长(s):"))
        self._min_dur_spin = QDoubleSpinBox()
        self._min_dur_spin.setRange(0.1, 60.0)
        self._min_dur_spin.setValue(0.5)
        self._min_dur_spin.setDecimals(1)
        self._min_dur_spin.setMaximumWidth(80)
        bottom_layout.addWidget(self._min_dur_spin)

        bottom_layout.addWidget(QLabel("最大时长(s):"))
        self._max_dur_spin = QDoubleSpinBox()
        self._max_dur_spin.setRange(0.5, 300.0)
        self._max_dur_spin.setValue(15.0)
        self._max_dur_spin.setDecimals(1)
        self._max_dur_spin.setMaximumWidth(80)
        bottom_layout.addWidget(self._max_dur_spin)

        apply_filter_btn = QPushButton("应用过滤")
        apply_filter_btn.setObjectName("secondaryBtn")
        bottom_layout.addWidget(apply_filter_btn)

        bottom_layout.addSpacing(30)

        # Split ratios
        bottom_layout.addWidget(QLabel("训练集 %:"))
        self._train_slider = QSlider(Qt.Horizontal)
        self._train_slider.setRange(50, 90)
        self._train_slider.setValue(80)
        self._train_slider.setMaximumWidth(150)
        bottom_layout.addWidget(self._train_slider)
        self._train_label = QLabel("80%")
        bottom_layout.addWidget(self._train_label)

        bottom_layout.addWidget(QLabel("验证集 %:"))
        self._val_slider = QSlider(Qt.Horizontal)
        self._val_slider.setRange(5, 30)
        self._val_slider.setValue(10)
        self._val_slider.setMaximumWidth(150)
        bottom_layout.addWidget(self._val_slider)
        self._val_label = QLabel("10%")
        bottom_layout.addWidget(self._val_label)

        bottom_layout.addWidget(QLabel("测试集 %:"))
        self._test_label = QLabel("10%")
        self._test_label.setStyleSheet("font-weight: bold;")
        bottom_layout.addWidget(self._test_label)

        apply_split_btn = QPushButton("重新划分")
        apply_split_btn.setObjectName("secondaryBtn")
        bottom_layout.addWidget(apply_split_btn)

        main_layout.addWidget(bottom_group)

        # === Connections ===
        self._import_btn.clicked.connect(lambda: self._import_files(recursive=False))
        self._import_recursive_btn.clicked.connect(lambda: self._import_files(recursive=True))
        export_jsonl_btn.clicked.connect(self._export_jsonl)
        import_xlsx_btn.clicked.connect(self._import_xlsx_labels)
        prepare_training_btn.clicked.connect(self._prepare_training_data)
        clear_btn.clicked.connect(self._clear_dataset)

        self._table.selectionModel().selectionChanged.connect(self._on_file_selected)
        self._play_btn.clicked.connect(self._play_audio)
        self._stop_btn.clicked.connect(self._stop_audio)

        apply_filter_btn.clicked.connect(self._apply_duration_filter)
        apply_split_btn.clicked.connect(self._apply_split)

        self._train_slider.valueChanged.connect(self._on_train_slider_changed)
        self._val_slider.valueChanged.connect(self._on_val_slider_changed)

    def _connect_signals(self):
        self._controller.import_finished.connect(self._on_import_finished)
        self._controller.dataset_updated.connect(self._refresh_table)

    # ================================================================
    # Slots
    # ================================================================

    @Slot()
    def _import_files(self, recursive: bool = False):
        directory = QFileDialog.getExistingDirectory(self, "选择音频文件夹")
        if not directory:
            return
        self._import_btn.setEnabled(False)
        count = self._controller.import_from_directory(directory, recursive=recursive)
        self._import_btn.setEnabled(True)

    @Slot(int)
    def _on_import_finished(self, count: int):
        self._file_count_label.setText(f"已导入: {self._controller.dataset.total_files} 个文件")

    @Slot()
    def _refresh_table(self):
        self._model.set_files(self._controller.files)
        self._file_count_label.setText(f"已导入: {self._controller.dataset.total_files} 个文件")

    @Slot()
    def _on_file_selected(self):
        indexes = self._table.selectionModel().selectedRows()
        if not indexes:
            self._play_btn.setEnabled(False)
            self._stop_btn.setEnabled(False)
            return
        info: AudioFileInfo = self._model.data(indexes[0], Qt.UserRole)
        if info and os.path.exists(info.path):
            self._preview_file_label.setText(info.filename)
            self._waveform.load_audio(info.path)
            self._current_preview_path = info.path
            self._play_btn.setEnabled(True)
            self._stop_btn.setEnabled(True)
            self._log.info("数据集", f"选中文件: {info.filename} ({info.duration:.1f}s)")
        else:
            self._play_btn.setEnabled(False)
            self._stop_btn.setEnabled(False)
            self._log.warning("数据集", f"无法加载: {info.path if info else 'N/A'}")

    @staticmethod
    def _read_audio_for_playback(path: str):
        """Read audio file with multi-backend fallback. Returns (audio, sr)."""
        import numpy as np

        # 1) soundfile
        try:
            import soundfile as sf
            audio, sr = sf.read(path)
            return audio, sr
        except Exception:
            pass
        # 2) scipy
        try:
            from scipy.io import wavfile
            sr, audio = wavfile.read(path)
            if audio.dtype == np.int16:
                audio = audio.astype(np.float32) / 32768.0
            elif audio.dtype == np.float64:
                audio = audio.astype(np.float32)
            return audio, sr
        except Exception:
            pass
        # 3) built-in wave
        import wave
        with wave.open(path, "rb") as wf:
            sr = wf.getframerate()
            n_frames = wf.getnframes()
            audio = np.frombuffer(wf.readframes(n_frames), dtype=np.int16)
            return audio.astype(np.float32) / 32768.0, sr

    @Slot()
    def _play_audio(self):
        self._log.info("数据集", f"播放按钮点击, path={self._current_preview_path}")
        if not self._current_preview_path:
            QMessageBox.information(self, "提示", "请先在文件列表中选择一个音频文件")
            return
        try:
            import sounddevice as sd
            import numpy as np
            audio, sr = self._read_audio_for_playback(self._current_preview_path)
            self._log.info("数据集", f"音频已加载: shape={audio.shape}, sr={sr}")
            if audio.ndim > 1:
                audio = audio.mean(axis=1)
            sd.stop()
            sd.play(audio, sr)
            self._log.info("数据集", "播放中...")
        except Exception as e:
            self._log.error("数据集", f"播放失败: {e}")
            QMessageBox.warning(self, "播放失败", f"无法播放音频:\n{e}")

    @Slot()
    def _stop_audio(self):
        self._log.info("数据集", "停止播放")
        try:
            import sounddevice as sd
            sd.stop()
        except Exception as e:
            self._log.warning("数据集", f"停止失败: {e}")

    @Slot()
    def _apply_duration_filter(self):
        mn = self._min_dur_spin.value()
        mx = self._max_dur_spin.value()
        self._controller.set_duration_filter(mn, mx)

    @Slot(int)
    def _on_train_slider_changed(self, value: int):
        self._train_label.setText(f"{value}%")
        val = self._val_slider.value()
        test = max(0, 100 - value - val)
        self._test_label.setText(f"{test}%")

    @Slot(int)
    def _on_val_slider_changed(self, value: int):
        self._val_label.setText(f"{value}%")
        train = self._train_slider.value()
        test = max(0, 100 - train - value)
        self._test_label.setText(f"{test}%")

    @Slot()
    def _apply_split(self):
        train = self._train_slider.value() / 100.0
        val = self._val_slider.value() / 100.0
        test = (100 - self._train_slider.value() - self._val_slider.value()) / 100.0
        if test < 0:
            QMessageBox.warning(self, "比例错误", "训练+验证比例不能超过100%")
            return
        self._controller.set_split_ratios(train, val, test)

    @Slot()
    def _import_xlsx_labels(self):
        """从 xlsx 文件导入文本标注，匹配到已导入的音频."""
        path, _ = QFileDialog.getOpenFileName(
            self, "选择标注文件(xlsx)", "",
            "Excel Files (*.xlsx);;All Files (*.*)")
        if not path:
            return

        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb.active
        except Exception as e:
            QMessageBox.critical(self, "读取失败", f"无法读取 xlsx 文件:\n{e}")
            return

        # Read all rows into a dict: filename -> {text, dialect, gender, age, region}
        label_map = {}
        for r in range(2, ws.max_row + 1):
            fname = str(ws.cell(r, 1).value or '').strip()
            if not fname:
                continue
            label_map[fname] = {
                'text': str(ws.cell(r, 6).value or '').strip(),
                'dialect': str(ws.cell(r, 7).value or '').strip(),
                'gender': str(ws.cell(r, 3).value or '').strip(),
                'age': str(ws.cell(r, 4).value or '').strip(),
                'region': str(ws.cell(r, 5).value or '').strip(),
            }

        # Match to imported files
        matched = 0
        files = self._controller.dataset.files
        for f in files:
            if f.filename in label_map:
                lbl = label_map[f.filename]
                f.text = lbl['text'] or f.text
                f.dialect = lbl['dialect'] or f.dialect
                f.gender = lbl['gender'] or f.gender
                f.age = lbl['age'] or f.age
                f.region = lbl['region'] or f.region
                matched += 1

        self._refresh_table()
        self._log.info("数据集",
            f"xlsx 标注导入完成: {matched}/{len(files)} 个文件已匹配")
        QMessageBox.information(self, "导入完成",
            f"从 xlsx 读取了 {len(label_map)} 条标注\n"
            f"匹配到 {matched} 个已导入的音频文件")

    @Slot()
    def _export_jsonl(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "导出 JSONL 元数据", "train.jsonl",
            "JSONL Files (*.jsonl);;All Files (*.*)"
        )
        if path:
            self._controller.export_metadata_jsonl(path)

    @Slot()
    def _prepare_training_data(self):
        """一键导出训练数据并通知训练面板."""
        valid_count = self._controller.dataset.statistics().get("valid", 0)
        if valid_count == 0:
            QMessageBox.warning(self, "无有效数据",
                "当前没有状态为 'valid' 的音频文件。\n\n"
                "请先导入音频文件夹，并确保文件时长在 min/max 范围内。")
            return

        # Default save path
        directory = QFileDialog.getExistingDirectory(
            self, "选择训练数据保存目录")
        if not directory:
            return

        path = os.path.join(directory, "train.jsonl")
        self._controller.export_metadata_jsonl(path)

        # Notify training panel via parent MainWindow
        from app.ui.main_window import MainWindow
        mw = self.window()
        if isinstance(mw, MainWindow) and "training" in mw._panels:
            tp = mw._panels["training"]
            tp._data_root = directory
            tp._data_dir_label.setText(directory)
            mw._tab_widget.setCurrentWidget(tp)
            self._log.info("数据集",
                f"训练数据已准备: {path} ({valid_count} 条)\n"
                f"请切换到 [训练配置] 面板，点击 [开始训练]")

        QMessageBox.information(self, "训练数据已准备",
            f"已导出 {valid_count} 条数据到:\n{path}\n\n"
            f"训练数据根目录已自动设置。\n"
            f"请切换到 [模型训练] 面板，点击 [开始训练] 启动训练。")

    @Slot()
    def _clear_dataset(self):
        reply = QMessageBox.question(
            self, "确认清空", "确定要清空所有数据集吗？此操作不可撤销。",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
        )
        if reply == QMessageBox.Yes:
            self._controller.clear()
            self._waveform.clear()
            self._preview_file_label.setText("选择文件以预览")

    # ================================================================
    # Public API
    # ================================================================

    @property
    def controller(self) -> DatasetController:
        return self._controller

    def get_dataset_config(self) -> dict:
        """Export current dataset config for project saving."""
        return self._controller.dataset.to_dict()
